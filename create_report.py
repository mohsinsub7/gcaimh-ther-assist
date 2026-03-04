"""
TherAssist Implementation Report Generator
Creates an Excel document covering all Phases 1-5 clinical feedback implementation.
"""
import sys
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
phase_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
phase_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
cell_font = Font(name='Calibri', size=10)
wrap_align = Alignment(wrap_text=True, vertical='top')
header_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ── Sheet 1: Implementation Report Data ──────────────────────────────
columns = [
    "Task ID",
    "Task Name\n(Meeting Reference)",
    "From Who",
    "Original Issue (GAP)",
    "What Was Added",
    "Where It Was Added",
    "What It Does Now",
    "How It Does It",
    "File Location",
    "Function / Code Reference",
    "Type"
]

col_widths = [9, 28, 14, 40, 42, 22, 42, 48, 38, 38, 16]

tasks = [
    # ── PHASE 1: SAFETY & EDGE CASE HANDLING ──
    {
        "phase": "Phase 1: Safety & Edge Case Handling",
        "id": "1A",
        "name": "Deterministic Safety\nKeyword Scanner\n(James: 'risk management\nfor edge cases')",
        "from": "James",
        "gap": "System relied 100% on the LLM to catch safety concerns. If the LLM hallucinated or missed a keyword, no safety alert would fire. There was zero deterministic fallback — the entire safety system was probabilistic.",
        "added": "New deterministic keyword scanner function that runs on EVERY transcript segment BEFORE the LLM call. Covers 5 clinical categories (suicidal ideation, violence/homicide, abuse/mandatory reporting, substance crisis, self-harm) with ~40 keywords total. Fixed 'meth' substring false positive by requiring word-boundary matching.",
        "where": "Backend: main.py (lines 720-788)\nConstants: constants.py (lines 35-173)",
        "does_now": "Every transcript segment is scanned by a hard-coded keyword detector BEFORE reaching the AI model. If any safety keyword matches, the system: (1) forces non-strict prompt mode, (2) escalates thinking_budget to maximum (24,576 tokens), (3) injects a safety flag into the prompt, and (4) creates a deterministic alert even if the LLM returns nothing.",
        "how": "Uses Python regex with word-boundary matching (\\b) to scan transcript text against 5 category-specific keyword lists. Each category has its own list (e.g., suicidal: 'kill myself', 'end my life', 'suicidal'; violence: 'hurt them', 'kill him', 'weapon'). Returns a SafetyScanResult dict with matched categories, keywords found, and severity. Runs in <1ms per segment.",
        "files": "main.py (detect_safety_keywords function, lines 720-788)\nconstants.py (SAFETY_KEYWORDS dict, lines 35-173)",
        "code": "detect_safety_keywords(text: str) -> dict\nSAFETY_KEYWORDS = { 'suicidal_ideation': [...], 'violence_homicide': [...], 'abuse_mandatory_reporting': [...], 'substance_crisis': [...], 'self_harm': [...] }",
        "type": "New Feature"
    },
    {
        "id": "1B",
        "name": "Expanded LLM Safety\nPrompts\n(James: 'homicide, violence,\nabuse, substances,\nmandatory reporting')",
        "from": "James",
        "gap": "LLM prompts only mentioned suicidal ideation. No instructions for violence toward others, child/elder abuse mandatory reporting, substance crisis, or crisis resource delivery. The AI had no guidance for 4 out of 5 safety categories.",
        "added": "Expanded both REALTIME_ANALYSIS_PROMPT and COMPREHENSIVE_ANALYSIS_PROMPT with explicit instructions for all 5 safety categories. Added SAFETY_CLINICAL_RESPONSES dictionary with per-category clinical protocols, crisis resources, immediate actions, and contraindications.",
        "where": "Backend constants:\nconstants.py (lines 78-173, 196-335)",
        "does_now": "The AI model receives detailed instructions for handling every safety category: suicidal ideation (C-SSRS protocol), violence/homicide (Tarasoff duty-to-warn), abuse (mandatory reporting statutes), substance crisis (SAMHSA protocols), and self-harm (safety planning). Each category includes specific crisis resources (988 Lifeline, 911, Childhelp National Hotline, SAMHSA Helpline).",
        "how": "SAFETY_CLINICAL_RESPONSES dict maps each category to: protocol_name (e.g., 'C-SSRS', 'Tarasoff'), crisis_resources (phone numbers/URLs), immediate_actions (clinical steps), and contraindications (what NOT to do). These are injected into LLM prompts when safety keywords are detected, providing structured clinical guidance.",
        "files": "constants.py (SAFETY_CLINICAL_RESPONSES dict, lines 196-335)\nconstants.py (REALTIME_ANALYSIS_PROMPT update, lines 78-173)",
        "code": "SAFETY_CLINICAL_RESPONSES = {\n  'suicidal_ideation': { protocol: 'C-SSRS', resources: ['988 Lifeline'] },\n  'violence_homicide': { protocol: 'Tarasoff', resources: ['911'] },\n  ...\n}",
        "type": "Upgrade"
    },
    {
        "id": "1C",
        "name": "Enhanced Safety\nAlert UI Banner\n(James: 'therapist needs\nclear visual alert with\ncrisis resources')",
        "from": "James",
        "gap": "Safety alerts displayed as regular orange chips in the guidance tab — identical to routine clinical suggestions. No special visual treatment, no crisis resource display, no acknowledgment flow. Easy to miss during a live therapy session.",
        "added": "Full-width red SafetyAlertBanner component at the top of the session view. Includes alert title, descriptive message, crisis phone numbers, and a mandatory 'Acknowledge' button. Banner persists until therapist clicks Acknowledge. Added crisis_resources field to Alert type and SafetyScanResult interface.",
        "where": "Frontend:\nNewTherSession.tsx (banner JSX + state)\ntypes.ts (new interface fields)",
        "does_now": "When a safety alert fires, a prominent red banner appears at the TOP of the therapist's session view — impossible to miss. It shows: the alert category (e.g., 'Suicidal Ideation Detected'), the clinical message, relevant crisis phone numbers (988, 911, Childhelp, SAMHSA), and an 'Acknowledge' button. The banner persists and cannot be dismissed until acknowledged.",
        "how": "React component renders conditionally when safetyAlert state is non-null. Uses MUI Alert component with severity='error' and custom red styling. crisis_resources array renders as clickable phone links. Acknowledge button sets safetyAlert state to null. State is lifted to NewTherSession parent component for persistence across tab switches.",
        "files": "NewTherSession.tsx (SafetyAlertBanner JSX, ~30 lines)\ntypes.ts (crisis_resources?: string[] field added to Alert type)",
        "code": "SafetyAlertBanner component\nNew fields: Alert.crisis_resources?: string[]\nSafetyScanResult interface in types.ts",
        "type": "New Feature"
    },
    {
        "id": "1E",
        "name": "Edge Case Safety\nTest Scripts\n(James + Team: 'validate\nsafety system for\nedge cases')",
        "from": "James + Team",
        "gap": "Only had 2 generic test transcripts (IPT→BA transitions). No safety-focused test scenarios. No way to systematically validate if the keyword scanner + LLM prompts fire correctly for crisis content across all 5 categories.",
        "added": "4 new test transcript JSON files simulating realistic therapy sessions that escalate into safety concerns. Each script is a JSON array of {speaker, text} entries loadable via the 'Upload Script' button in the session view.",
        "where": "Test scripts directory:\ngcaimh-ther-assist-DEV/test-scripts/\n(4 new JSON files)",
        "does_now": "Clinical team can load any of these scripts into a test session to verify the safety system fires correctly. Each script simulates a realistic therapeutic conversation that gradually escalates into a crisis scenario, testing both the keyword scanner and LLM safety detection.",
        "how": "JSON arrays of {speaker: 'therapist'|'patient', text: '...'} entries. Loaded via Upload Script button → fed through transcript pipeline → triggers detect_safety_keywords() → sends to LLM with safety prompts → should produce safety alerts with crisis resources.\n\nScripts:\n1. safety_suicidal_ideation.json — Active SI with plan + means access\n2. safety_abuse_disclosure.json — Child abuse/DV mandatory reporting\n3. safety_violence_homicide.json — Homicidal ideation toward specific person\n4. safety_substance_crisis.json — Overdose/withdrawal crisis",
        "files": "test-scripts/safety_suicidal_ideation.json\ntest-scripts/safety_abuse_disclosure.json\ntest-scripts/safety_violence_homicide.json\ntest-scripts/safety_substance_crisis.json",
        "code": "Each file: JSON array of ~20-30 speaker/text entries\nFormat: [{\"speaker\": \"therapist\", \"text\": \"...\"}, {\"speaker\": \"patient\", \"text\": \"...\"}]",
        "type": "New Feature"
    },
    # ── PHASE 2: THERAPY MODALITY RESTRUCTURING ──
    {
        "phase": "Phase 2: Therapy Modality Restructuring",
        "id": "2A",
        "name": "RAG Routing Hierarchy\nRestructure (CBT Umbrella)\n(Amy: 'CBT should be\numbrella containing\nEBT+CBT+BA')",
        "from": "Amy",
        "gap": "MODALITY_RAG_MAP mapped each modality to a single RAG datastore. CBT only pulled from cbt-corpus, not the related ba-corpus or ebt-corpus. No multi-corpus support — selecting one modality queried only one datastore, missing relevant literature from related approaches.",
        "added": "Changed MODALITY_RAG_MAP values from single tools to LISTS of tools. CBT/BA umbrella now pulls from ebt-corpus + cbt-corpus + ba-corpus simultaneously. Updated get_rag_tools_for_session() to iterate over tool lists with deduplication. Updated comprehensive analysis fallback.",
        "where": "Backend:\nmain.py (lines 367-416,\nlogging blocks, fallback tools)",
        "does_now": "When therapist selects CBT or BA, the system queries 3 RAG datastores simultaneously (EBT + CBT + BA). DBT queries EBT + DBT. IPT queries EBT + IPT. This ensures comprehensive literature coverage across related therapeutic approaches.\n\nFull mapping:\nCBT → [ebt, cbt, ba]\nBA → [ebt, cbt, ba]\nExposure → [ebt, cbt]\nDBT → [ebt, dbt]\nIPT → [ebt, ipt]\nACT → [ebt, cbt]\nGeneral → [ebt, cbt]",
        "how": "MODALITY_RAG_MAP dict values changed from single tool references to Python lists. get_rag_tools_for_session() iterates the list, deduplicates (using set), and returns combined tool list. Logging dynamically builds corpus name strings from the active tool list. Comprehensive analysis fallback includes BA corpus for broader coverage.",
        "files": "main.py (MODALITY_RAG_MAP dict, lines 367-390)\nmain.py (get_rag_tools_for_session(), lines 391-416)",
        "code": "MODALITY_RAG_MAP = {\n  'cbt': [ebt_tool, cbt_tool, ba_tool],\n  'ba': [ebt_tool, cbt_tool, ba_tool],\n  'dbt': [ebt_tool, dbt_tool],\n  ...\n}\nget_rag_tools_for_session(session_type) -> list[Tool]",
        "type": "Modification"
    },
    {
        "id": "2B",
        "name": "Therapist Modality\nSelection Dialog\n(Team: 'therapist should\nexplicitly choose therapy\ntype at session start')",
        "from": "Team Consensus",
        "gap": "Session started recording immediately when therapist clicked 'Start Session'. Modality was auto-inferred from the patient's focusTopics field — therapist had no input. If focusTopics was wrong or generic, the wrong RAG corpus loaded for the entire session.",
        "added": "Session Setup Dialog (MUI Dialog) that gates the 'Start Session' flow. Before recording begins, therapist sees 7 radio options (CBT, DBT, IPT, ACT, Exposure, BA, General) with clinical descriptions. Selected modality locks sessionContext.session_type and sets therapist_selected_modality flag.",
        "where": "Frontend:\nNewTherSession.tsx (dialog state,\nMODALITY_OPTIONS array,\nhandleConfirmSessionStart)\ntypes.ts (new field)",
        "does_now": "When therapist clicks 'Start Session', a modal dialog appears BEFORE recording begins. The dialog presents 7 therapy modalities with brief descriptions (e.g., 'CBT — Cognitive Behavioral Therapy: Structured approach focusing on thoughts, feelings, and behaviors'). Therapist selects one, clicks 'Confirm', and THEN recording starts. A blue 'Therapist Selected' badge displays throughout the session.",
        "how": "MUI Dialog component with RadioGroup. MODALITY_OPTIONS constant array holds {value, label, description} for each modality. handleConfirmSessionStart() sets sessionContext.session_type to selected value, sets therapist_selected_modality: true, closes dialog, and initiates recording. The flag prevents AI from overriding the therapist's explicit choice.",
        "files": "NewTherSession.tsx (SessionSetupDialog, ~80 lines)\nNewTherSession.tsx (MODALITY_OPTIONS array)\ntypes.ts (therapist_selected_modality: boolean)",
        "code": "MODALITY_OPTIONS = [{value: 'cbt', label: 'CBT', description: '...'}, ...]\nhandleConfirmSessionStart(selectedModality: string)\ntherapist_selected_modality field in SessionContext type",
        "type": "New Feature"
    },
    {
        "id": "2C",
        "name": "Disable Mid-Session\nModality Auto-Switching\n(Team: 'avoid switch in\nsession between\ntherapy types')",
        "from": "Team Consensus",
        "gap": "When the AI detected a different modality at >=70% confidence mid-session, it silently changed sessionContext.session_type — switching RAG corpus mid-conversation. Therapist might not notice. Clinically disruptive: RAG citations suddenly change from CBT to DBT literature without therapist consent.",
        "added": "Removed the auto-switch code block. AI-detected modality now stored in sessionMetrics.detected_modality for display only. Three-state display system: 'Therapist Selected' (blue badge), 'AI suggests: X' (amber chip), or 'Gathering data...' (grey chip).",
        "where": "Frontend:\nNewTherSession.tsx\n(replaced auto-switch block,\nupdated chip rendering)",
        "does_now": "The system NEVER automatically switches therapy modality mid-session. The therapist's explicit selection (from Task 2B) is locked for the entire session. The AI's modality detection still runs but is display-only: shown as an amber 'AI suggests: X (Y%)' chip that the therapist can note for future sessions. Three visual states: (1) Blue badge = therapist selected, (2) Amber chip = AI suggestion differs, (3) Grey chip = gathering data.",
        "how": "Deleted the conditional block (old lines 493-508) that called setSessionContext when AI confidence >= 70%. AI detection result now writes to sessionMetrics.detected_modality (read-only). Chip rendering logic uses conditional: if therapist_selected_modality → show blue badge; else if detected != selected AND words >= threshold → show amber chip; else → show grey 'Gathering data' chip.",
        "files": "NewTherSession.tsx (removed auto-switch block)\nNewTherSession.tsx (updated modality chip rendering)",
        "code": "Removed: if (confidence >= 0.7) setSessionContext({session_type: detected})\nAdded: sessionMetrics.detected_modality = detected (display-only)\nThree-state chip rendering logic",
        "type": "Modification"
    },
    {
        "id": "2E",
        "name": "Minimum Transcript\nThreshold for Modality\nDetection\n(Team: 'suggestion should\nbe based on enough\ntranscript')",
        "from": "Team Consensus",
        "gap": "AI would suggest a modality after just 1-2 transcript entries (~30 words). Not clinically meaningful — insufficient therapeutic content for reliable modality detection. Resulted in noisy, low-confidence suggestions that changed frequently in early session.",
        "added": "Word count tracking (totalWordCountRef) across all transcript entries. MIN_WORDS_FOR_MODALITY_SUGGESTION = 200 (~2-3 minutes of conversation). Frontend suppresses modality suggestion display until threshold met. Backend has matching constant.",
        "where": "Frontend: NewTherSession.tsx\n(word counting + threshold guard)\nBackend: constants.py\n(MIN_WORDS constant)",
        "does_now": "The system waits until at least 200 words of transcript have been accumulated (~2-3 minutes of conversation) before displaying any AI modality suggestion. Until then, a grey 'Gathering data...' chip is shown. This ensures the AI has enough therapeutic content to make a clinically meaningful modality assessment.",
        "how": "totalWordCountRef (React useRef) increments on every transcript entry from 3 sources: finalized microphone transcripts, test transcript entries, and transcript playback. Split by whitespace, count words, accumulate. Modality chip rendering checks: if totalWordCountRef.current < MIN_WORDS_FOR_MODALITY_SUGGESTION → show 'Gathering data...' chip. Resets to 0 on new session start.",
        "files": "NewTherSession.tsx (totalWordCountRef, word counting logic)\nconstants.py (MIN_WORDS_FOR_MODALITY_DETECTION = 200)",
        "code": "totalWordCountRef = useRef<number>(0)\nMIN_WORDS_FOR_MODALITY_SUGGESTION = 200\nMIN_WORDS_FOR_MODALITY_DETECTION = 200 (backend)",
        "type": "New Feature"
    },
    # ── PHASE 3: EVALUATION & DATA WIRING ──
    {
        "phase": "Phase 3: Evaluation & Data Pipeline",
        "id": "3B",
        "name": "Wire TherSummary to\nLive Session Data\n(Team: 'evaluate real\nsession summaries, not\nhardcoded demo data')",
        "from": "Team",
        "gap": "TherSummary.tsx rendered with hardcoded defaults: 'John Doe', 41 min duration, 18 techniques. The component had prop interfaces for real data but App.tsx never passed it. When a session ended, the summary always showed fake placeholder data regardless of what the AI generated.",
        "added": "Full data pipeline from session end to summary view. Added lastSessionSummary state in App.tsx. Expanded handleSessionSaved callback. Mapped SessionSummary fields to TherSummary props: progress_indicators, areas_for_improvement, follow_up_recommendations, homework_assignments (with title/description/reference mapping), risk_assessment.level with color-coded display.",
        "where": "Frontend:\nApp.tsx (state, callback, rendering)\nNewTherSession.tsx\n(onSessionSaved call + prop type)",
        "does_now": "When a session ends and the AI generates a comprehensive summary, ALL real data flows through to the TherSummary view: actual progress indicators, real areas for improvement, AI-generated follow-up recommendations, homework assignments with clinical references, and color-coded risk assessment level. Falls back gracefully to defaults if no AI summary is available.",
        "how": "App.tsx: lastSessionSummary useState stores SessionSummary from completed session. handleSessionSaved callback receives fullSummary parameter from NewTherSession. Mapping layer converts backend SessionSummary shape to TherSummary prop shape (e.g., homework_assignments[].title → homeworkTitle, .description → homeworkDesc, .reference → homeworkRef). Risk level maps to color: 'low' → green, 'moderate' → amber, 'high' → red.",
        "files": "App.tsx (lastSessionSummary state, handleSessionSaved, rendering)\nNewTherSession.tsx (onSessionSaved fullSummary parameter)",
        "code": "lastSessionSummary = useState<SessionSummary | null>(null)\nhandleSessionSaved(sessionId, patientId, fullSummary?: SessionSummary)\nProp mapping: summary.homework_assignments → TherSummary homework props",
        "type": "Modification"
    },
    # ── PHASE 5: SAFETY RAG + PATIENT INTERFACE ──
    {
        "phase": "Phase 5: Safety RAG & Patient Interface",
        "id": "5A",
        "name": "Safety & Crisis RAG\nDocument Corpus\n(Clinical team: evidence-\nbased safety protocols\nfor AI grounding)",
        "from": "Clinical Team",
        "gap": "The AI safety system had no evidence-based clinical literature to ground its responses. Safety alerts and crisis guidance were generated from LLM training data only — not from authoritative clinical protocols like C-SSRS, SAMHSA guidelines, or mandatory reporting standards.",
        "added": "9 open-access clinical PDF documents downloaded and ingested into a new 'safety-crisis' Vertex AI Search datastore. New SAFETY_RAG_TOOL always included in every session's tool list (not modality-dependent). Setup script created for reproducible deployment.",
        "where": "GCP: safety-crisis datastore\nGCS: brk-prj-salvador-dura-bern-sbx-safety-crisis bucket\nBackend: main.py (SAFETY_RAG_TOOL)\nSetup: setup_services/rag/",
        "does_now": "Every therapy session — regardless of modality — has access to evidence-based safety and crisis literature. When the AI detects safety concerns, it can ground its recommendations in authoritative clinical protocols (C-SSRS screening criteria, SAMHSA crisis care guidelines, Stanley-Brown safety planning, mandatory reporting statutes) rather than relying solely on training data.",
        "how": "9 PDFs uploaded to GCS bucket → imported into Vertex AI Search datastore 'safety-crisis'. SAFETY_RAG_TOOL defined using Vertex AI genai Tool class pointing to the datastore. Tool is appended to every session's tool list in get_rag_tools_for_session() regardless of selected modality. AI model can cite specific passages from these documents in its safety guidance.\n\nDocuments:\n- C-SSRS Baseline Screening & Full Baseline\n- 988 Lifeline Risk Assessment Standards & Safety Policy\n- SAMHSA National Guidelines for Crisis Care\n- SAMHSA SAFE-T Suicide Assessment\n- SAMHSA TIP50 (Suicidal Thoughts + Substance Abuse)\n- Stanley-Brown Safety Plan\n- Child Welfare Mandatory Reporting Statutes",
        "files": "setup_services/rag/corpus_safety/ (9 PDF files)\nsetup_services/rag/setup_safety_datastore.py\nmain.py (SAFETY_RAG_TOOL definition + inclusion)",
        "code": "SAFETY_RAG_TOOL = Tool.from_retrieval(\n  retrieval=Retrieval(source=VertexAISearch(\n    datastore='safety-crisis')))\nAlways included: tools.append(SAFETY_RAG_TOOL)",
        "type": "New Feature"
    },
    {
        "id": "5B",
        "name": "Session History for\nModality Recommendation\n(Team: 'base therapy type\non previous sessions')",
        "from": "Team Consensus",
        "gap": "The modality selection dialog (Task 2B) pre-selects based on patient's focusTopics field, but has no knowledge of what therapy modalities were used in previous sessions. Therapist must remember past session types from memory.",
        "added": "BLOCKED — Requires Firebase/Firestore read permissions that are not currently available in the dev environment. Implementation plan ready: backend endpoint get_patient_session_history to fetch last 3 sessions, frontend useEffect to display 'Previous sessions: CBT (2), BA (1)' in setup dialog.",
        "where": "N/A — Blocked on\nFirebase permissions",
        "does_now": "Currently NOT implemented. The modality selection dialog pre-selects based on patient focusTopics (existing behavior). Once Firebase access is granted, will show session history recommendations above the modality radio buttons.",
        "how": "Planned: Backend queries Firestore for patient's last 3 session documents, extracts session_type field, counts frequency. Returns {modalities: [{type: 'cbt', count: 2, last_date: '...'}, ...]}. Frontend fetches on dialog open, displays as recommendation chip above radio buttons, pre-selects most frequent modality.",
        "files": "Planned:\nmain.py (get_patient_session_history endpoint)\nNewTherSession.tsx (useEffect fetch in dialog)",
        "code": "Planned:\nget_patient_session_history(patient_id) -> dict\nFrontend: useEffect(() => fetchHistory(patientId), [patientId])",
        "type": "Blocked (Planned Feature)"
    },
    {
        "id": "5C",
        "name": "Patient Post-Session\nInterface\n(Team: 'patient-facing\nsummary view')",
        "from": "Team / Tarek",
        "gap": "No patient-facing view existed. The TherSummary component showed clinical-internal data (risk levels, technique counts, clinical citations, therapist notes) that is inappropriate for patient viewing. Patients had no way to review homework or session takeaways.",
        "added": "New PatientSummary.tsx component with simplified, patient-friendly display. New route in App.tsx. New backend endpoint get_patient_summary returning filtered data subset (no risk assessment, no clinical citations). 'View Patient Summary' button added to TherSummary.",
        "where": "Frontend:\nPatientSummary.tsx (NEW)\nApp.tsx (new route + view)\nTherSummary.tsx (button)\nBackend:\nmain.py (get_patient_summary)",
        "does_now": "After a session, therapist can click 'View Patient Summary' to see what the patient would see. The patient view shows: session date/duration, homework assignments with interactive checkboxes, progress highlights (framed positively), follow-up recommendations in plain language, and next session reminders. It strips out: risk levels, technique counts, clinical citations, and therapist-only notes.",
        "how": "PatientSummary.tsx reuses TherSummary data model but filters and reformats. Homework displayed as cards with checkboxes (task + rationale, no manual references). Progress indicators use encouraging language. Backend get_patient_summary action filters SessionSummary to remove clinical-internal fields before returning. Route '/patient-summary' in App.tsx renders the component with passed data.",
        "files": "frontend/components/PatientSummary.tsx (NEW, ~200 lines)\nfrontend/components/App.tsx (route + navigation)\nfrontend/components/TherSummary.tsx (button)\nbackend/therapy-analysis-function/main.py (get_patient_summary)",
        "code": "PatientSummary component (React functional component)\nget_patient_summary(session_id) -> filtered dict\nApp.tsx: case 'patientSummary': <PatientSummary ... />\nTherSummary: onViewPatientSummary callback",
        "type": "New Feature"
    }
]

# ── Sheet 2: RAG & Corpus Reference Data ──────────────────────────────
rag_data = [
    ["RAG Datastore Name", "Corpus Directory", "Content Description", "# Documents", "Modalities Using It"],
    ["ebt-corpus", "corpus_ebt", "Evidence-Based Treatment manuals (PE, CBT-Social Phobia, Deliberate Practice)", "~5", "ALL modalities (base corpus)"],
    ["cbt-corpus", "corpus_cbt", "31 CBT randomized controlled trials and clinical studies", "31", "CBT, BA, Exposure, ACT, General"],
    ["ba-corpus", "corpus_ba", "11 Behavioral Activation treatment studies", "11", "CBT, BA"],
    ["dbt-corpus", "corpus_dbt", "6 DBT randomized controlled trials and systematic reviews", "6", "DBT"],
    ["ipt-corpus", "corpus_ipt", "10 IPT randomized controlled trials and meta-analyses", "10", "IPT"],
    ["transcript-patterns", "corpus_transcripts", "Beck Institute sessions + ThousandVoicesOfTrauma therapy dataset", "~10", "Pattern matching (all)"],
    ["safety-crisis", "corpus_safety", "Clinical safety protocols, crisis guidelines, mandatory reporting (see below)", "9", "ALL sessions (always included)"],
]

safety_docs = [
    ["Document Name (PDF)", "Source Organization", "Clinical Use"],
    ["C-SSRS_Baseline_Screening.pdf", "Columbia University (cssrs.columbia.edu)", "Suicide risk screening — baseline questions"],
    ["C-SSRS_Full_Baseline.pdf", "Columbia University (cssrs.columbia.edu)", "Full Columbia Suicide Severity Rating Scale"],
    ["988_Lifeline_Suicide_Risk_Assessment_Standards.pdf", "988 Suicide & Crisis Lifeline", "Risk assessment best practices for crisis counselors"],
    ["988_Lifeline_Suicide_Safety_Policy.pdf", "988 Suicide & Crisis Lifeline", "Safety policy for suicide crisis intervention"],
    ["SAMHSA_National_Guidelines_Crisis_Care.pdf", "SAMHSA (US Gov, public domain)", "National guidelines for behavioral health crisis care"],
    ["SAMHSA_SAFE-T_Suicide_Assessment.pdf", "SAMHSA (US Gov, public domain)", "Suicide Assessment Five-step Evaluation and Triage"],
    ["SAMHSA_TIP50_Suicidal_Thoughts_Substance_Abuse.pdf", "SAMHSA (US Gov, public domain)", "Addressing suicidal thoughts in substance abuse treatment"],
    ["Stanley-Brown_Safety_Plan.pdf", "Stanley & Brown (open clinical use)", "Safety Planning Intervention template and guide"],
    ["ChildWelfare_Mandatory_Reporting_Statutes.pdf", "Children's Bureau (childwelfare.gov, US Gov)", "State-by-state mandatory reporting statutes for child abuse"],
]

# ── Build Workbook ────────────────────────────────────────────────────
wb = Workbook()

# ── SHEET 1: Implementation Report ──
ws1 = wb.active
ws1.title = "Implementation Report"

# Title row
ws1.merge_cells('A1:K1')
title_cell = ws1['A1']
title_cell.value = "TherAssist — Clinical Feedback Implementation Report (Phases 1-5)"
title_cell.font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

# Subtitle row
ws1.merge_cells('A2:K2')
subtitle = ws1['A2']
subtitle.value = "Project: brk-prj-salvador-dura-bern-sbx | Environment: DEV (gcaimh-ther-assist-DEV/) | Date: February 20, 2026"
subtitle.font = Font(name='Calibri', size=10, italic=True, color='666666')
subtitle.alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 20

# Header row (row 4)
for col_idx, header in enumerate(columns, 1):
    cell = ws1.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws1.row_dimensions[4].height = 40

# Set column widths
for col_idx, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# Data rows
row = 5
current_phase = None
for task in tasks:
    # Phase separator row
    if "phase" in task and task["phase"] != current_phase:
        current_phase = task["phase"]
        ws1.merge_cells(f'A{row}:K{row}')
        phase_cell = ws1.cell(row=row, column=1, value=current_phase)
        phase_cell.font = phase_font
        phase_cell.fill = phase_fill
        phase_cell.alignment = Alignment(vertical='center')
        phase_cell.border = thin_border
        ws1.row_dimensions[row].height = 25
        row += 1

    # Data row
    values = [
        task["id"], task["name"], task["from"], task["gap"],
        task["added"], task["where"], task["does_now"], task["how"],
        task["files"], task["code"], task["type"]
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws1.cell(row=row, column=col_idx, value=val)
        cell.font = cell_font
        cell.alignment = wrap_align
        cell.border = thin_border

    # Type column color
    type_cell = ws1.cell(row=row, column=11)
    if "New Feature" in task["type"]:
        type_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    elif "Upgrade" in task["type"]:
        type_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    elif "Modification" in task["type"]:
        type_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    elif "Blocked" in task["type"]:
        type_cell.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

    ws1.row_dimensions[row].height = 120
    row += 1

# Freeze panes
ws1.freeze_panes = 'A5'

# ── SHEET 2: RAG & Corpus Reference ──
ws2 = wb.create_sheet("RAG & Corpus Reference")

# Title
ws2.merge_cells('A1:E1')
ws2['A1'].value = "RAG Datastores & Corpus Reference"
ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
ws2['A1'].alignment = Alignment(horizontal='center')
ws2.row_dimensions[1].height = 30

# RAG table
for row_idx, row_data in enumerate(rag_data, 3):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = wrap_align
        if row_idx == 3:  # Header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        else:
            cell.font = cell_font

rag_widths = [22, 20, 55, 14, 35]
for col_idx, w in enumerate(rag_widths, 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

# Safety documents section
safety_start = 3 + len(rag_data) + 2
ws2.merge_cells(f'A{safety_start}:C{safety_start}')
ws2[f'A{safety_start}'].value = "Safety RAG Corpus Documents (safety-crisis datastore)"
ws2[f'A{safety_start}'].font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
ws2.row_dimensions[safety_start].height = 25

for row_idx, row_data in enumerate(safety_docs, safety_start + 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = wrap_align
        if row_idx == safety_start + 1:  # Header
            cell.font = header_font
            cell.fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
            cell.alignment = header_align
        else:
            cell.font = cell_font

ws2.freeze_panes = 'A4'

# ── Save ──────────────────────────────────────────────────────────────
output_path = r"C:\Users\mohsi\OneDrive\Documents\Personal Projects\AI Mental Health Model\GMAIC Model\gcaimh-ther-assist-DEV\TherAssist_Implementation_Report.xlsx"
wb.save(output_path)
print(f"SUCCESS: Report saved to {output_path}")
print(f"Sheet 1: Implementation Report ({len(tasks)} tasks across 4 phases)")
print(f"Sheet 2: RAG & Corpus Reference ({len(rag_data)-1} datastores, {len(safety_docs)-1} safety docs)")
